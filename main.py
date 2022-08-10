from openpyxl import Workbook
import pandas as pd
import os
import time

from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import decimal
from pathlib import Path
from datetime import date
import glob


def format_number(num):
    try:
        dec = decimal.Decimal(num)
    except:
        return num
    val = dec
    return val

def latest_edited_file_glob(pathToDir, **kwargs):
    pathTest = os.path.dirname(__file__)

    if pathToDir == 'ori':
        ori_file_path = os.path.join(pathTest, 'data', 'ori', 'CLASSCHED.csv')
        if not os.path.isfile(ori_file_path):
            print('ori is empty. assume you want to compare to most recent colored diff')
            latest_file = 'empty'
        else:
            list_of_files = glob.glob(
                os.path.join(pathTest, "data", "ori", '*'))  # * means all if need specific format then *.csv
            latest_file = max(list_of_files, key=os.path.getctime)
            print("glob latest_file:" + latest_file)

    elif pathToDir == 'changed':
        list_of_files = glob.glob(
            os.path.join(pathTest, "data", "changed", '*'))  # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        print("glob latest_file:" + latest_file)

    else:
        list_of_files = glob.glob(
            os.path.join(pathTest, "color_diffs", "Spring 2022", '*'))  # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        print("glob latest_file:" + latest_file)

    return latest_file


def read_csv(filename):
    # latest_edited_file_glob('test') # to get date
    path = os.path.dirname(__file__)
    print('File name :    ', os.path.basename(__file__))
    print('Directory Name:     ', path)

    if filename == 'ori':
        #latest_edited_file = max([f for f in os.scandir(os.path.join(path, "data", "ori"))], key=lambda x: x.stat().st_mtime).name
        #pathFile = os.path.join(path, "data", "ori", latest_edited_file)
        pathFile = latest_edited_file_glob('ori')

        try:
            modified = os.path.getmtime(pathFile)
            year, month, day, hour = time.localtime(modified)[:-5]
            file_date = str(str(year) + '_' + str(month) + '_' + str(day))
            print("Original Date:", file_date)
            print('pathFile is:', pathFile)
        except:
            file_date = time.ctime(os.path.getmtime(pathFile))

    elif filename == 'new':
        pathFile = latest_edited_file_glob('changed')
        if pathFile == 'empty':
            print('')
        try:
            modified = os.path.getmtime(pathFile)
            year, month, day, hour = time.localtime(modified)[:-5]
            file_date = str(str(year)+'_'+str(month)+'_'+str(day))
            print ("New Date:", file_date)
            file_date = str(str(year) + '_' + str(month) + '_' + str(day))
            print("New Date:", file_date)
        except:
            file_date = time.ctime(os.path.getmtime(pathFile))
        print('pathFile is:', pathFile)

    df = pd.read_csv(pathFile, header=None)

    return df, file_date

def duplicate_entry_merger(df):
    df = df.applymap(str)
    df = df.groupby([df.columns[0], df.columns[7]]).agg(" ; ".join).reset_index()
    return df


if __name__ == '__main__':
    # pathFile = latest_edited_file_glob('changed')
    today = date.today()

    df_modified_input_file, new_date = read_csv('new')
    df_modified_file = df_modified_input_file.loc[df_modified_input_file.iloc[:, 14] == "CHEM"]

    for i in range(df_modified_file.shape[0]):
        for j in range(df_modified_file.shape[1]):
            df_modified_file.iloc[i, j] = format_number(df_modified_file.iloc[i, j])

    filtered_columns = [0, 4, 7, 15, 16, 17, 18, 19, 21, 23, 25, 27, 34, 35, 36, 38, 51, 53, 57, 66, 67, 81, 82, 83, 90,
                        91,
                        97, 110, 112]
    df_filtered_mod = df_modified_file.iloc[:]

    if len(df_modified_file.columns) > 140:  # avoids applying filters to short files
        df_filtered_mod = df_modified_file.iloc[:, filtered_columns]


    modified_term = df_filtered_mod.iloc[0, 1]
    print('modified_term is:', modified_term)

    pathTest = os.path.dirname(__file__)
    pathFile = latest_edited_file_glob('ori')
    if pathFile == 'empty':
        pathFile = os.path.join(pathTest, "color_diffs", f"{modified_term}")
        list_of_files = glob.glob(
            os.path.join(pathFile, '*.csv'))  # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        pathFile = latest_file
        print("glob latest_file:" + latest_file)
        # modified = os.path.getmtime(pathFile)
        # year, month, day, hour = time.localtime(modified)[:-5]
        # ori_date = str(str(year) + '_' + str(month) + '_' + str(day))

    # ori_file_path = os.path.join(pathTest, 'data', 'ori', 'CLASSCHED.csv')
    # if not os.path.isfile(ori_file_path):
    #     print('ori is empty. assume you want to compare to most recent colored diff')
    #     pathFile = os.path.join(pathTest, "color_diffs", f"{modified_term}", '*')
        modified = os.path.getmtime(pathFile)
        year, month, day, hour = time.localtime(modified)[:-5]
        ori_date = str(str(year) + '_' + str(month) + '_' + str(day))
        # ori_date = time.ctime(os.path.getmtime(pathFile))
    else:
        modified = os.path.getmtime(pathFile)
        year, month, day, hour = time.localtime(modified)[:-5]
        ori_date = str(str(year) + '_' + str(month) + '_' + str(day))
        print("Original Date:", ori_date)
        print('pathFile is:', pathFile)

    df_original_file = pd.read_csv(pathFile, header=None)

    # df_original_file, ori_date = read_csv('ori')
    df_original_file = df_original_file.loc[df_original_file.iloc[:, 14] == "CHEM"]

    for i in range(df_original_file.shape[0]):
        for j in range(df_original_file.shape[1]):
            df_original_file.iloc[i, j] = format_number(df_original_file.iloc[i, j])


    df_filtered_ori = df_original_file.iloc[:]

    if len(df_modified_file.columns) > 140:  # avoids applying filters to short files
       df_filtered_mod = df_modified_file.iloc[:, filtered_columns]


    print('83 ', df_filtered_mod[83])
    df_instructor = df_filtered_mod.iloc[:,[3,6,8,17,23]]
    print('df_instructor ', df_instructor)
    print(df_instructor.shape)
    # ff = pd.pivot_table(df_instructor, values=['83'], index=['53', '83'], columns=['53'], aggfunc=np.sum, fill_value=0)


    term = df_filtered_mod.iloc[0, 1]
    print('term' + term)

    Path("color_diffs/" + term).mkdir(parents=True, exist_ok=True)

    # ----------------
    df_filtered_ori = duplicate_entry_merger(df_filtered_ori)
    df_filtered_mod = duplicate_entry_merger(df_filtered_mod)
    # ----------------
    df_modified_input_file.to_csv(f'color_diffs/{modified_term}/{new_date}_full_out.csv', sep=',', index=False, header=False)

    df_changes = pd.concat([df_filtered_ori.assign(type='original'), df_filtered_mod.assign(type='modified')])
    df_changes = df_changes.drop_duplicates(keep=False, subset=df_changes.columns.difference(['type']))

    df_modified_file.to_csv(f'color_diffs/{modified_term}/{ori_date}_{new_date}_diff_out.csv', sep=',', index=False, header=False)

    df_changes_arranged = df_changes

    df_changes_arranged = df_changes_arranged.sort_values(
        by=[df_changes_arranged.columns[0], df_changes_arranged.columns[7]], ascending=True)
    # df_changes_arranged.to_csv(f'color_diffs/{modified_term}/{ori_date}_{new_date}_diff_out_arranged.csv', sep=',', index=False,
    #                            header=False)


    j = 0
    i = 0
    c = 0
    rows_count = df_changes_arranged.shape[0]
    cols_count = df_changes_arranged.shape[1]
    # print('rows_count:' + str(rows_count))

    changed_id = []
    unique_id = []

    unique_id_from_ori = []
    unique_id_from_mod = []

    red_color = 'background-color: red'
    green_color = 'background-color: lightgreen'

    for i in range(rows_count - 1):
        if ((df_changes_arranged.iloc[i, 0] == df_changes_arranged.iloc[i + 1, 0]) and (
                df_changes_arranged.iloc[i, 7] == df_changes_arranged.iloc[i + 1, 7])):
            # print("same rows detected" + str(i))
            j = 0

            for c in range(cols_count):
                if c < 29:
                    if df_changes_arranged.iloc[i, c] != df_changes_arranged.iloc[i + 1, c]:
                        # print('--- different column' + str(c))
                        changed_id.append([i, c])
                        # changed_id.append([i + 1, c])

        else:
            j = j + 1
            if j >= 2:
                # print('unique row detected' + str(i))
                unique_id.append([i + 1, 1])
                j = 0

            elif i == rows_count - 2:
                unique_id.append([i + 2, 1])
                j = 0

            else:
                print("different rows detected" + str(i))

    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df_changes_arranged, index=False, header=False):
        ws.append(r)

    fill_pattern_modified_content = PatternFill(patternType='solid', fgColor='ffaba6')
    fill_pattern_modified_content_darker = PatternFill(patternType='solid', fgColor='f03629')
    fill_pattern_unique_content = PatternFill(patternType='solid', fgColor='61bdff')

    fill_pattern_from_original = PatternFill(patternType='solid', fgColor='FFA500')
    fill_pattern_from_modified = PatternFill(patternType='solid', fgColor='66CDAA')

    for l in range(len(df_changes_arranged)):
        if ws.cell(l + 1, 30).value == "original":
            # print("original")
            for p in range(cols_count):
                ws.cell(l + 1, p + 1).fill = fill_pattern_from_original
        elif ws.cell(l + 1, 30).value == "modified":
            # print("modified")
            for p in range(cols_count):
                ws.cell(l + 1, p + 1).fill = fill_pattern_from_modified

    for i in range(len(changed_id)):
        ws.cell(changed_id[i][0] + 1, changed_id[i][1] + 1).fill = fill_pattern_modified_content
        ws.cell(changed_id[i][0] + 2, changed_id[i][1] + 1).fill = fill_pattern_modified_content_darker

    wb.save("color_diffs/" + term + "/" + today.isoformat() + "_output.xlsx")
